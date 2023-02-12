import glob
import os
import zipfile
import openpyxl as op
import fnmatch
import pathlib
import shutil


# 폴더별 파일명 받아 엑셀파일에 작성하는 함수
def getFileName(trg_path):
    # 여러 폴더가 있는 폴더경로를 입력받아 list화 한다.
    folderlist = os.listdir(trg_path)

    # openpyxl Workbook 생성
    wb = op.Workbook()

    # 새로 생성한 Workbook의 활성화 시트를 ws로 정의
    ws = wb.active

    i = 2

    current_folder = trg_path  # 현재 폴더 위치
    filelist = os.listdir(current_folder)  # 현재 폴더 위치의 파일들을 리스트화

    # 현재 폴더 위치의 리스트화 된 파일을 다시 for문을 통해 접근
    for fname in filelist:
        ws.cell(row=i, column=1).value = current_folder + "/"  # i행 1열에 현재 폴더 경로 입력
        ws.cell(row=i, column=2).value = fname  # i행 2열에 파일명(변경전) 입력
        i = i + 1

    ws.cell(row=1, column=1).value = "파일경로"
    ws.cell(row=1, column=2).value = "파일명(변경전)"
    ws.cell(row=1, column=3).value = "파일명(변경후)"

    # 위에서 작성 된 엑셀파일을 filelist.xlsx라는 이름으로 저장
    wb.save(os.path.join(trg_path, "filelist.xlsx"))


def excelRead(filepath: str) -> tuple:
    # filelist.xlsx 경로를 받아 workbook 객체 생성
    wb = op.load_workbook(filepath)

    # 활성화 된 시트 ws로 정의
    ws = wb.active

    # 리스트 컴프리헨션 구문을 통해 각 열의 데이터를 리스트화하기
    folderpath = [r[0].value for r in ws]  # 폴더 경로 리스트화
    file_before = [r[1].value for r in ws]  # 변경전 파일명 리스트화
    file_after = [r[2].value for r in ws]  # 변경후 파일명 리스트화

    len_num = len(folderpath)  # for문을 위한 열 개수 구하기(folderpath 개수)
    datalist = []  # 빈 리스트 생성
    for i in range(1, len_num):
        temp_tuple = (folderpath[i], file_before[i], file_after[i])  # 임시 튜플 생성(각 행 데이터)
        datalist.append(temp_tuple)  # 위 튜플을 datalist에 추가

    return datalist  # datalist 리턴


def fileRename(datalist : list):
    for data in datalist:
        print(data[1]+"의 파일명을"+data[2]+"로 변경합니다.")
        #data[0] : 폴더경로, data[1] : 변경전 파일명, data[2] : 변경 후 파일명
        os.rename(data[0]+data[1], data[0]+data[2])


# 파일명을 읽어와서 파일명의 분류 부분을 중복없이 리스트화
def categoryList(trg_path: str) -> list:
    # 파일명 끝자리가 _XXX (숫자 세자리) 로 끝나는 파일 탐색
    file_list = []
    for filename in os.listdir(trg_path):
        if fnmatch.fnmatch(filename, '*_[0-9][0-9][0-9].*'):
            file_list.append(filename)

    category = []  # 분류 데이터 저장을 위해 빈 리스트 생성

    for file in file_list:
        temp_list = file.split("_")  # 파일명중 "_"로 분리하여 리스트화
        category.append(temp_list[-2])  # 리스트의 -2 인덱싱 데이터를 category에 추가

    temp_set = set(category)  # 중복을 제거하기 위해 set 사용
    result = list(temp_set)  # 중복 제거 후 다시 리스트화
    return result  # 결과 리턴

def makeFolder(clean_path : str, categorylist : list):
    for category in categorylist:
        new_folder = pathlib.Path(os.path.join(clean_path, category))
        new_folder.mkdir(parents=True, exist_ok=True)


# 파일을 폴더 분류에 맞게 이동
def moveFile(clean_path, trg_path, categorylist):
    folderlist = os.listdir(clean_path)  # 이동시킬 경로에 생성된 분류별 폴더 리스트화
    filelist = os.listdir(trg_path)  # 이동시킬 파일명들을 리스트화
    categorydict = {}  # 빈 딕셔너리 생성

    # 파일명에 대한 폴더명을 딕셔너리로 저장
    for file in filelist:

        # 파일명 규칙에 맞지 않으면 '기타' 폴더로 분류
        try:
            temp_list = file.split("_")
            assert temp_list[-2] in categorylist  # 카테고리가 맞지 않으면 에러 발생

            categorydict[file] = temp_list[-2]  # {'파일명' : '분류'} 형태의 딕셔너리 생성
        except:
            categorydict[file] = '기타'

    # 딕셔너리 정보 활용하여 파일 이동
    for key, value in categorydict.items():
        shutil.copy(trg_path + "/" + key, clean_path + "/" + value)



if __name__ == "__main__":
    # 정리 대상 폴더 경로 지정
    trg_path = './fuzzy_folder_1'
    clean_path = './clean_folder'

    # 압축 파일 확인
    zipf_path = []
    for filename in glob.glob(os.path.join(trg_path, "**/*.zip"), recursive=True):
        zipf_path.append(filename)
        print(zipf_path)

    # 압축 파일 해제
    for zipf in zipf_path:
        with zipfile.ZipFile(zipf) as myzip:
            zipinfo = myzip.infolist()
            for info in zipinfo:
                decode_name = info.filename.encode('cp437').decode('euc-kr')  # 한글 깨짐 방지
                info.filename = os.path.join(trg_path, decode_name)
                myzip.extract(info)

    # 함수 실행
    getFileName(trg_path)

    # 함수 실행
    rename_list = excelRead(os.path.join(trg_path, "filelist.xlsx"))
    print(rename_list)

    # 함수 실행
    fileRename(rename_list)

    # 함수 실행
    categorylist = categoryList(trg_path) + ['기타']
    print(categorylist)

    # 함수 실행
    makeFolder(clean_path, categorylist)

    # 함수 실행
    moveFile(clean_path, trg_path, categorylist)